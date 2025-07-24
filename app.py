import pandas as pd
import os
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import load_workbook

# 🔧 Set to your new root-level folder
BASE_DIR = "/storage/emulated/0/AlignmentApp/"
EXCEL_FILE = os.path.join(BASE_DIR, "master_vehicle_database_alignment.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

SHEET_MASTER = "master vehicle list"
SHEET_ALIGNMENT = "Daily Alignment"
BRANCH_SHEETS = ["Obajana", "Ibese"]
ACTIONS = ["Vehicle to be presented at Tyre-Bay", "Presented"]
REMARKS = ["Vehicle not yet Presented", "Present but sent to workshop for repair", "Presented and aligned"]

def clean_license(lic):
    lic = str(lic).strip().upper()
    if lic.endswith("C") or (lic.endswith("T") and not lic.endswith("THT")):
        return lic[:-1]
    elif lic.endswith("CH"):
        return lic[:-2]
    elif lic.endswith("CNG"):
        return lic[:-3]
    return lic

def clean_vehicle_id(veh_id):
    veh_id = str(veh_id).strip().upper()
    if veh_id.endswith("T") and not veh_id.endswith("THT"):
        return veh_id[:-1]
    elif veh_id.endswith("CH"):
        return veh_id[:-2]
    elif veh_id.endswith("CNG"):
        return veh_id[:-3]
    return veh_id

def is_body_id(veh_id):
    veh_id = str(veh_id).strip().upper()
    return (
        veh_id.startswith(("DT", "TIP", "TP", "GDT", "DS")) or
        (veh_id.endswith("T") and not veh_id.endswith("THT")) or
        veh_id.endswith("CH") or
        veh_id.endswith("CNG")
    )

def generate_branch_output(master_df, branch_df):
    master_df['License'] = master_df['License'].astype(str).str.upper().str.strip()
    master_df['Vehicle#'] = master_df['Vehicle#'].astype(str).str.upper().str.strip()

    master_df['is_body'] = master_df['Vehicle#'].apply(is_body_id)
    master_df['CleanLicense'] = master_df['License'].apply(clean_license)
    master_df['CleanVehicle'] = master_df.apply(lambda row: clean_vehicle_id(row['Vehicle#']) if row['is_body'] else row['Vehicle#'], axis=1)

    branch_df = branch_df.copy()
    branch_df['Vehicle#'] = branch_df['Vehicle#'].astype(str).str.upper().str.strip()

    matched_branch = pd.merge(branch_df, master_df[['Vehicle#', 'License']], on='Vehicle#', how='left')
    matched_branch['CleanLicense'] = matched_branch['License'].apply(clean_license)
    unique_licenses = matched_branch['CleanLicense'].dropna().unique()

    tractors = master_df[~master_df['is_body']][['CleanLicense', 'CleanVehicle', 'Route']].drop_duplicates()
    bodies = master_df[master_df['is_body']][['CleanLicense', 'CleanVehicle']].drop_duplicates()

    final_rows = []
    for lic in unique_licenses:
        tractor = tractors[tractors['CleanLicense'] == lic]
        body = bodies[bodies['CleanLicense'] == lic]

        tractor_id = tractor['CleanVehicle'].values[0] if not tractor.empty else ""
        body_id = body['CleanVehicle'].values[0] if not body.empty else ""
        route = tractor['Route'].values[0] if not tractor.empty else ""

        final_rows.append({
            "Tractor": tractor_id,
            "Body": body_id,
            "License": lic,
            "Route": route,
            "Action": ACTIONS[0],
            "Remark": REMARKS[0],
            "Date Aligned": ""
        })

    df_final = pd.DataFrame(final_rows)

    df_final["has_both"] = df_final["Tractor"].ne("") & df_final["Body"].ne("")
    df_final["only_tractor"] = df_final["Tractor"].ne("") & df_final["Body"].eq("")
    df_final["only_body"] = df_final["Tractor"].eq("") & df_final["Body"].ne("")

    df_sorted = pd.concat([
        df_final[df_final["has_both"]],
        df_final[df_final["only_tractor"]],
        df_final[df_final["only_body"]]
    ])

    return df_sorted.drop(columns=["has_both", "only_tractor", "only_body"]).reset_index(drop=True)

def save_to_excel_with_dropdowns(results_by_branch, output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, df in results_by_branch.items():
        ws = wb.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        ws.freeze_panes = "A2"

        action_col = 'E'
        remark_col = 'F'
        last_row = ws.max_row

        dv_action = DataValidation(type="list", formula1=f'"{",".join(ACTIONS)}"', showDropDown=True)
        dv_remark = DataValidation(type="list", formula1=f'"{",".join(REMARKS)}"', showDropDown=True)
        ws.add_data_validation(dv_action)
        ws.add_data_validation(dv_remark)
        dv_action.add(f"{action_col}2:{action_col}{last_row}")
        dv_remark.add(f"{remark_col}2:{remark_col}{last_row}")

    wb.save(output_path)

def main():
    print("🚚 Vehicle Daily Report Generator (Output Folder Version)")
    print(f"📥 Reading Excel file: {EXCEL_FILE}")

    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Error: File '{EXCEL_FILE}' not found")
        input("✔ Press Enter to exit...")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_file = os.path.join(OUTPUT_DIR, f"vehicle_alignment_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    try:
        excel_data = pd.read_excel(EXCEL_FILE, sheet_name=None)
        master_df = excel_data.get(SHEET_MASTER)

        if master_df is None:
            raise ValueError(f"Worksheet named '{SHEET_MASTER}' not found")

        results = {}
        for branch in BRANCH_SHEETS:
            branch_df = excel_data.get(branch)
            if branch_df is None:
                print(f"⚠️ Warning: Sheet '{branch}' not found. Skipping.")
                continue

            result_df = generate_branch_output(master_df.copy(), branch_df)
            results[branch] = result_df

        if results:
            save_to_excel_with_dropdowns(results, output_file)
            print(f"\n✅ Output saved to: {output_file}")
        else:
            print("❌ No valid branch sheets processed.")

    except Exception as e:
        print(f"❌ An error occurred: {e}")

    input("✔ Press Enter to exit...")

if __name__ == "__main__":
    main()